from rest_framework import viewsets, permissions, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters import rest_framework as django_filters 
from .models import Category, Component
from .serializers import CategorySerializer, ComponentSerializer
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from apps.interactions.models import StockNotification

class ComponentFilter(django_filters.FilterSet):
    min_price = django_filters.NumberFilter(field_name="price", lookup_expr='gte')
    max_price = django_filters.NumberFilter(field_name="price", lookup_expr='lte')
    
    mpn = django_filters.CharFilter(field_name="mpn", lookup_expr='iexact')
    
    valor = django_filters.CharFilter(field_name="technical_specs__valor", lookup_expr='icontains')
    voltaje = django_filters.CharFilter(field_name="technical_specs__voltaje", lookup_expr='icontains')
    tolerancia = django_filters.CharFilter(field_name="technical_specs__tolerancia", lookup_expr='icontains')
    encapsulado = django_filters.CharFilter(field_name="technical_specs__encapsulado", lookup_expr='icontains')
    montaje = django_filters.CharFilter(field_name="technical_specs__montaje", lookup_expr='iexact')

    class Meta:
        model = Component
        fields = ['category', 'is_available', 'store']

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.AllowAny] 

class ComponentViewSet(viewsets.ModelViewSet):
    serializer_class = ComponentSerializer
    filter_backends = [django_filters.DjangoFilterBackend, filters.SearchFilter]
    filterset_class = ComponentFilter
    search_fields = ['name', 'mpn', 'description', 'technical_specs__valor', 'technical_specs__encapsulado']

    def get_queryset(self):
        queryset = Component.objects.all()
        is_management = self.request.query_params.get('manage', None)
        
        if is_management and self.request.user.is_authenticated:
            if not self.request.user.is_staff:
                return queryset.filter(store__owner=self.request.user)
        return queryset

    def perform_create(self, serializer):
        from .models import Store 
        user_store = Store.objects.get(owner=self.request.user)
        serializer.save(store=user_store)

    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'recommendations', 'price_comparison']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]
    
    @action(detail=False, methods=['get'])
    def recommendations(self, request):
        if request.user.is_authenticated:
            user_interests = Component.objects.filter(in_wishlists__user=request.user).values_list('category', flat=True)
            if user_interests.exists():
                recommended = Component.objects.filter(category__in=user_interests).exclude(in_wishlists__user=request.user).distinct()[:5]
            else:
                recommended = Component.objects.all().order_by('?')[:5]
        else:
            recommended = Component.objects.all().order_by('?')[:5]
            
        serializer = self.get_serializer(recommended, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def toggle_notification(self, request, pk=None):
        """Activa o desactiva el seguimiento de stock para un componente"""
        component = self.get_object()
        notification, created = StockNotification.objects.get_or_create(
            user=request.user, 
            component=component
        )
        
        if not created:
            # Si ya existía, lo borramos (toggle off)
            notification.delete()
            return Response({'is_notifying': False, 'message': 'Seguimiento desactivado'})
        
        # Si se creó nuevo (toggle on)
        return Response({'is_notifying': True, 'message': 'Te avisaremos cuando haya poco stock'})

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def low_stock_alerts(self, request):
        """Devuelve solo los componentes seguidos por el usuario que tienen stock bajo"""
        # Filtramos componentes que tengan una StockNotification activa para este usuario
        # y que además cumplan la condición de stock (ej. < 10)
        queryset = Component.objects.filter(
            stocknotification__user=request.user,
            stocknotification__is_active=True,
            stock__lt=10
        ).distinct().order_by('stock')
        
        low_stock = queryset.filter(stock__gt=0)
        out_of_stock = queryset.filter(stock__lte=0)
        
        return Response({
            "low_stock": ComponentSerializer(low_stock, many=True, context={'request': request}).data,
            "out_of_stock": ComponentSerializer(out_of_stock, many=True, context={'request': request}).data,
            "total_alerts": queryset.count()
        })
        
    @action(detail=True, methods=['get'])
    def price_comparison(self, request, pk=None):
        component = self.get_object()
        
        if not component.mpn:
            return Response([])

        # Buscamos coincidencias
        comparisons = Component.objects.filter(
            mpn__iexact=component.mpn
        ).exclude(id=component.id)

        # IMPORTANTE: Usar el serializador para que devuelva store_name e imagen
        serializer = self.get_serializer(comparisons, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def download_excel(self, request): 
        queryset = self.get_queryset()
        if not request.user.is_staff:
            queryset = queryset.filter(store__owner=request.user)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario de Tienda"

        # 1. Definir Estilos
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Morado elegante
        header_font = Font(color="FFFFFF", bold=True, size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(bottom=Side(style='thin', color="DDDDDD"))

        # 2. Encabezados
        headers = ['Componente', 'MPN', 'Precio', 'Stock', 'Estatus']
        ws.append(headers)

        # Aplicar estilos a los encabezados
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # 3. Agregar Datos con formato condicional básico
        for item in queryset:
            status_text = "Agotado" if item.stock <= 0 else "Disponible"
            row_data = [item.name, item.mpn, item.price, item.stock, status_text]
            ws.append(row_data)
            
            # Estilo para la última celda (Estatus)
            last_cell = ws.cell(row=ws.max_row, column=5)
            if status_text == "Agotado":
                last_cell.font = Font(color="EF4444", bold=True) # Rojo
            else:
                last_cell.font = Font(color="10B981", bold=True) # Verde

        # 4. AUTO-AJUSTAR columnas (El toque final para que se vea bien)
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 5

        # 5. Respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_inventario.xlsx'
        wb.save(response)
        return response
